"""
Celery tasks for background data ingestion.
"""
import logging
from celery import shared_task

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def ingest_customer_data(self, file_path: str):
    """
    Ingest customer data from Excel file into the database.
    """
    try:
        import openpyxl
        from .models import Customer

        logger.info(f"Starting customer data ingestion from {file_path}")
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        created_count = 0
        updated_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            row_data = dict(zip(headers, row))

            customer_id = row_data.get('Customer ID') or row_data.get('customer_id')
            first_name = row_data.get('First Name') or row_data.get('first_name', '')
            last_name = row_data.get('Last Name') or row_data.get('last_name', '')
            phone_number = row_data.get('Phone Number') or row_data.get('phone_number', 0)
            monthly_salary = row_data.get('Monthly Salary') or row_data.get('monthly_salary', 0)
            approved_limit = row_data.get('Approved Limit') or row_data.get('approved_limit', 0)
            current_debt = row_data.get('Current Debt') or row_data.get('current_debt', 0)

            if customer_id:
                _, created = Customer.objects.update_or_create(
                    id=customer_id,
                    defaults={
                        'first_name': str(first_name) if first_name else '',
                        'last_name': str(last_name) if last_name else '',
                        'phone_number': int(phone_number) if phone_number else 0,
                        'monthly_salary': float(monthly_salary) if monthly_salary else 0,
                        'approved_limit': float(approved_limit) if approved_limit else 0,
                        'current_debt': float(current_debt) if current_debt else 0,
                    }
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1

        logger.info(f"Customer ingestion complete: {created_count} created, {updated_count} updated")
        return {'created': created_count, 'updated': updated_count}

    except Exception as exc:
        logger.error(f"Customer ingestion failed: {exc}")
        raise self.retry(exc=exc, countdown=5)


@shared_task(bind=True, max_retries=3)
def ingest_loan_data(self, file_path: str):
    """
    Ingest loan data from Excel file into the database.
    """
    try:
        import openpyxl
        from datetime import datetime
        from .models import Customer, Loan

        logger.info(f"Starting loan data ingestion from {file_path}")
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        created_count = 0
        skipped_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            row_data = dict(zip(headers, row))

            customer_id = row_data.get('Customer ID') or row_data.get('customer_id')
            loan_id = row_data.get('Loan ID') or row_data.get('loan_id')
            loan_amount = row_data.get('Loan Amount') or row_data.get('loan_amount', 0)
            tenure = row_data.get('Tenure') or row_data.get('tenure', 0)
            interest_rate = row_data.get('Interest Rate') or row_data.get('interest_rate', 0)
            monthly_repayment = row_data.get('Monthly payment') or row_data.get('monthly_repayment', 0)
            emis_paid_on_time = row_data.get('EMIs paid on Time') or row_data.get('emis_paid_on_time', 0)
            start_date = row_data.get('Date of Approval') or row_data.get('start_date')
            end_date = row_data.get('End Date') or row_data.get('end_date')

            try:
                customer = Customer.objects.get(id=customer_id)
            except Customer.DoesNotExist:
                logger.warning(f"Customer {customer_id} not found, skipping loan {loan_id}")
                skipped_count += 1
                continue

            # Parse dates
            def parse_date(d):
                if d is None:
                    return None
                if isinstance(d, datetime):
                    return d.date()
                if hasattr(d, 'date'):
                    return d.date()
                try:
                    return datetime.strptime(str(d), '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    try:
                        return datetime.strptime(str(d), '%d/%m/%Y').date()
                    except (ValueError, TypeError):
                        return None

            _, created = Loan.objects.update_or_create(
                id=loan_id,
                defaults={
                    'customer': customer,
                    'loan_amount': float(loan_amount) if loan_amount else 0,
                    'tenure': int(tenure) if tenure else 0,
                    'interest_rate': float(interest_rate) if interest_rate else 0,
                    'monthly_repayment': float(monthly_repayment) if monthly_repayment else 0,
                    'emis_paid_on_time': int(emis_paid_on_time) if emis_paid_on_time else 0,
                    'start_date': parse_date(start_date),
                    'end_date': parse_date(end_date),
                }
            )
            if created:
                created_count += 1

        logger.info(f"Loan ingestion complete: {created_count} created, {skipped_count} skipped")
        return {'created': created_count, 'skipped': skipped_count}

    except Exception as exc:
        logger.error(f"Loan ingestion failed: {exc}")
        raise self.retry(exc=exc, countdown=5)